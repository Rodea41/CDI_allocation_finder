
import PyPDF2
import re, csv, os
from csv import DictWriter
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font



#! This is the directory where the inventory file is stored 
#inventory_directory = "C:\\Users\\Chris R\\Desktop\\Python Projects\\Mini_Projects\\cdi_reader"  #! Use this at home
##inventory_directory = "C:\\Users\\crodea\\Desktop\\\Mini_Projects\\cdi_reader"  #! Use this at work 
inventory_directory = "C:\\Users\\crodea\\Desktop\\OneDrive\\OneDrive - US Cold Storage\\Python\\CDI_allocations_project" #! One drive folder
#print(file_directory)

#! This is the inventory file
inventory_file = "inventory.csv"


#!This is the directory that holds the allocation PDF's 
path = "C:\\Users\\crodea\\Desktop\\test_allocations"  #! This is the path for work desktop
#path = "C:\\Users\\Chris R\\Desktop\\Python Projects\\Mini_Projects\\cdi_reader\\test_allocations"   #! This is the path at home 


#! Open Up file and get the contents of the first page 
# def get_text(cdi_pdf):
#     pdfFileObj = open('CDI1.pdf', 'rb')
#     pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
#     pageObj = pdfReader.getPage(0)
#     text = pageObj.extractText()
#     return text

#! This will get the total number of pages for each pdf. 
def get_page_cnt(pdf): #! example of possible input ==> ['CDI0.pdf', 'CDI1.pdf', 'CDI2.pdf']
    invoice = pdf
    pdfFileObj = open(f'{invoice}', 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    pages = pdfReader.getNumPages()
    return pages


#! This function will find the order Number located in the PDF and store it in a variable to be used later as the output csv file name
def getOrderNumber(pdf):
    pattern = re.compile("SO000\\d\\d\\d\\d\\d\\d", re.IGNORECASE)
    match = pattern.findall(pdf)
    return match[0] 

#! Gets the pallet ID from the CDI allocation PDF file that is opened 
def getPalletId(cdi_pallets):
    pattern = re.compile("\\d\\d\\d\\d-\\d\\d\\d-\\d\\d-[a-zA-Z]\\d\\d-[a-zA-Z]", re.IGNORECASE)
    match = pattern.findall(cdi_pallets)
    #print(match)
    return match #! Returns a list of pallet ID's ==> ['3822-153-08-S04-B', '6622-153-13-S02-F', '6622-153-13-S03-C', '6622-153-13-S04-B']

#! Gets the quantity that is requested on the CDI order manifest. *NOT* the quantity in inventory
def getQTY(input):
    pattern = re.compile("(\d+\.0000\s+CS)", re.MULTILINE)
    match = pattern.findall(input) #! Returns ['84.0000\n\n\nCS', '84.0000\n\n\nCS', '84.0000\nCS']
    qty = []
    for items in match:
        #print(items)
        strip_new_line = items.split("\n") #! Returns ['84.0000', '', '', 'CS']
        add_CS = strip_new_line[0] + " CS" #! Returns 84.0000 CS
        
        qty.append(add_CS)
    
    return qty

#! Renames all the cdi pdfs in the folder 'test_allocations' so that we can loop through them 
def rename_files():
    import os

    #! Finds the path to the folder that holds the allocation PDF's 
    #path = "C:\\Users\\crodea\\Desktop\\test_allocations"  #! This is the path at work 
    #path = "C:\\Users\\Chris R\\Desktop\\Python Projects\\Mini_Projects\\cdi_reader\\test_allocations"   #! This is the path at home 

    #! Changes the directory to the path with the files 
    os.chdir(path)

    #! Iterates through files and renames them CDI0,CDI1,CDI2,etc. 
    for count, f in enumerate(os.listdir()):
        f_name, f_ext = os.path.splitext(f)
        f_name = "CDI" + str(count)

        new_name = f'{f_name}{f_ext}'
        os.rename(f, new_name)

    #! Create a combined list of file names that we can interate through
    file_list = []
    files = os.listdir()    
    for items in files:
        #print(items)
        file_list.append(items)

    return file_list

#! Opens the inventory excel sheet (8-3-1 report in WMS). looks up each pallet ID that is in the PID list
#! found in the getPalletId() function
def read_from_inventory_csv(number, qty):
    #csv_file = csv.reader(open("inventory.csv", "r"), delimiter=",") #! need to open the file within the for loop, or else it will not restart the search from top of file
    row_info = []
    os.chdir(inventory_directory) #! Change back from the pdf folder.

    # print("Pallet Count is : " + str(len(number[1])))
    # print("Qty Count is : " + str(len(qty)))
    Num_of_pallets = len(number[1])
    Num_of_qty = len(qty)
    # print(Num_of_pallets)
    # print(Num_of_qty)

    for items in range(len(number)):
        csv_file = csv.reader(open(f"{inventory_file}", "r"), delimiter=",")
        palletID = number[items]
        qty_requested = qty[items] #! Gets the qty from the pdf. If it is blank....

        #* NEED TO CREATE A CONDITIONAL STATEMENT FOR WHEN PALLET ID AND QTY DONT MATCH . 
        #print(number)
        add_comma = "'" + str(palletID) #! Turns 3822-150-05-S01-D into '3822-150-05-S01-D . Because the inventory csv adds a comma in front of PID
        #print(add_comma)
        for row in csv_file:
            #print(number[items])
            if add_comma == row[13]:
                #print(row)
                info = {
                    "Product Code": " ".join(row[2].split()),
                    "Lot Number": " ".join(row[1].split()), #! Strips the extra spaces 
                    "Quantity Requested From CDI": qty_requested, 
                    "Quantity in INV": " ".join(row[8].split()),
                    "Batch ID": palletID,
                    "Pallet ID Matched": " ".join(row[13].split()),
                    "License Number": " ".join(row[12].split())
                }

                row_info.append(info) #! Adds to the row_info list, which the function will return later. 
                break
            else:
                """Need to create a template for product that doesn't have a match """
                #print(str(palletID) + " No match " + str(row[13]))
                continue

        continue

    return row_info


#! This func is used to get all the lots that are being requested. We will use this list that is returned to check against what is being requested
def get_entire_lot(list_of_pallets):
    all_of_lot = []     #* Returns ['90786', '90786', '34924', '34924', '33964', '33964', '33964', '97884', '5630', '92', '92', '92', '1127', '1127', '1127', '1127', '1127', '1127', '1127'] . Shows duplicates
    for items in list_of_pallets:
        #print(items.get('Lot Number'))
        lot_num = items.get('Lot Number')
        all_of_lot.append(lot_num)
    #print(all_of_lot)

    rm_dups = set(all_of_lot) #* Returns {'91619', '91445', '97562', '91446'} . Removes all duplicates in list 'all_of_lot'
    #print(rm_dups)
    bk_to_list = list(rm_dups)

    return bk_to_list



#! This function strips all the spaces in the inventory excel sheet
def trim_all_columns(df):
    """
    Trim whitespace from ends of each value across all series in dataframe
    """
    trim_strings = lambda x: x.strip() if isinstance(x, str) else x
    return df.applymap(trim_strings)


#! Takes the list returned from func 'get_entire_lot' and looks get everything in that lot in the inventory csv
def read_lots_from_csv(lots):
    os.chdir(inventory_directory) #! Change back from the pdf folder.
    
    lot_info =[]
    
    #for items in range(len(lots)):
    for items in lots:
        #print(items)
        df = pd.read_csv(f"{inventory_file}")
        df = trim_all_columns(df)
        gk = df.groupby(['Lot','Pallet Id'])
        gk = df.loc[df["Lot"] == f"{items}"]
        to_dictionary = gk.to_dict('records')
    #gk = df.groupby(['Pallet Id', 'Lot'])
    #print(gk.get_group('91619'))
        filter_dictionary = []
        for row in to_dictionary:
            #print(row)
            relevant_info = {
                        "Lot": row.get('Lot'),  
                        "Pallet ID Matched": row.get('Pallet Id'),
                        "License Number": row.get('License  ')
                        }
            filter_dictionary.append(relevant_info)
        
        lot_info.append(filter_dictionary)
        #lot_info.append(to_dictionary)
    #print(lot_info)
    return lot_info

    #print(lots[0])
    # match = df.set_index(['Lot'])
    #match1 = match[match['Lot'] == '91445']
    #match = df.loc[df['Lot'] == '3592']
    #print(match)
    
    #by_lot = df.groupby(["Lot"]())

def format_and_style(xlsx_file):

    test = xlsx_file.strip('.xlsx')
    print(test)
    print(type(test))
    
    wb = load_workbook(xlsx_file)
    #wb = load_workbook(f'{orderNumber}.xlsx')
    ws = wb.active

    #! Setting the column width
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20

    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")

    #! Get all the licenses in column 'G' . Add them to a list called 'lice' (short for licenses)
    for i in range(1, 8): # Go through every other row:
        cell = ws.cell(row=1, column=i)
        cell.font = Font(b=True, italic=True)
        cell.fill = PatternFill("solid", fgColor="d4f4ad")
        cell.border = Border(top=double, left=thin, right=thin, bottom=double)

    for i in range(len(ws['A'])):
        cell = ws.cell(row=i+1, column=1)
        if cell.value == 'Lot':
            cell.font = Font(b=True, italic=True)
            cell.fill = PatternFill("solid", fgColor="d4f4ad")
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            cell = ws.cell(row=i+1, column=2)
            cell.font = Font(b=True, italic=True)
            cell.fill = PatternFill("solid", fgColor="d4f4ad")
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            cell = ws.cell(row=i+1, column=3)
            cell.font = Font(b=True, italic=True)
            cell.fill = PatternFill("solid", fgColor="d4f4ad")
            cell.border = Border(top=double, left=thin, right=thin, bottom=double)      
            
            
            # print(cell)
            # print(cell.value)



    lice = []
    for c in ws['G']:
        if c.value != None and c.value != 'License Number':
            #print(c.value)
            c.font = Font(b=True, italic=True)
            lice.append(c.value)

    #! Search through column 'C' for rows that have any of the values in the 'lice' list
    for f in ws['C']:
        #print(f.value)
        if f.value in lice:
            #print(f.value)
            #print("---")
            f.font = Font(b=True, color="FF0000", italic=True)
            f.fill = PatternFill("solid", fgColor="88f208")


    #wb.save("styled.xlsx")
    wb.save(f'{test}.xlsx')


def main():
    pdfs = rename_files()
    print(pdfs)

    for pdf in pdfs:
        os.chdir(path) #! Need to change back to the folder where the pdf's are because the 'read_from_inventory_csv function changes the path to 
        print(pdf)
        invoice = pdf
        pdfFileObj = open(f'{invoice}', 'rb')
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        pages = pdfReader.getNumPages()
        pageObj = pdfReader.getPage(0)
        text = pageObj.extractText()    

        orderNumber = getOrderNumber(text)

        #print(pages)
        myPalletList = [] #! Stores all the pallet ID's from each page of CDI pdf. 

        for page in range(pages):
            pageObj = pdfReader.getPage(page)
            text = pageObj.extractText()
            PID = getPalletId(text) #! Gets all the Pallet ID on Manifest
            QTY_REQUESTED = getQTY(text) #! Gets all the qty requested on Manifest
            if len(PID) == 0: #! If any page is empty or has no pallet ID, program will move on to next page
                continue
            else:
                for palletID in PID: #! Adds the pallet id's 1 by 1 instead of as a whole so multiple pages, wont build a list of lists
                    myPalletList.append(palletID)
                #print(myPalletList)
        
        matched_pallets = read_from_inventory_csv(myPalletList, QTY_REQUESTED)
        lots = get_entire_lot(matched_pallets)
        #print(lots)
        #!!! CURRENTLY WORKING ON THIS PART . LOOKING HOW TO GROUP LOOKUP ALL LOTS AVAILABLE.
        all_pallets_in_lots = read_lots_from_csv(lots)

        with open(f'{orderNumber}.csv','w', newline='') as outfile:
                writer = DictWriter(outfile, ('Product Code','Lot Number','Quantity Requested From CDI','Quantity in INV', 'Batch ID', 'Pallet ID Matched', 'License Number'))
                writer.writeheader()
                writer.writerows(matched_pallets)
                outfile.write("\n")
        
        #! This writes all the palletID's that are in the lot, this will help show if the order is using the whole lot or if product is left over. 
        with open(f'{orderNumber}.csv','a', newline='') as outfile:
            for items in range(len(all_pallets_in_lots)):
                    writer = DictWriter(outfile, ('Lot','Pallet ID Matched','License Number'))
                    writer.writeheader()
                    writer.writerows(all_pallets_in_lots[items])
                    outfile.write("\n")
    
        
        #! This converts the csv file to an xlsx file that can be formatted with color. 
        wb = openpyxl.Workbook()
        ws = wb.active

        with open(f'{orderNumber}.csv') as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)

        #! Saves the file without formatting 
        wb.save(f'{orderNumber}.xlsx')
        #wb.save("C:\\Users\\crodea\\Desktop\\Completed_CDI\\f'{orderNumber}.xlsx'")

        #! This adds the border, font color, etc to the .xlsx sheet that is created
        format_and_style(f'{orderNumber}.xlsx')



main()
